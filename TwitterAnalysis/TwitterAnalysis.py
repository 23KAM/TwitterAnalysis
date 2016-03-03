"""
Author: KAM Wright
Description: This code pull information on tweets related to screennames and outputs into an Excel file
Version: 1

"""
       
import tweepy
import openpyxl as xl
import datetime

consumer_key = 'MC7lINcv3bYIcDroF3WhXjZPa'
consumer_secret = 'pYTtJJcodQnq8lFyDzSqRo2aOdRvnQxJkXijuomKjJHRAHMBPm'
access_key = '22926126-blDZEHzX7yAHKqMqFnemG1ngbV826kglw4rfnyGMa'
access_secret = 'MRSOWlwRE5lqduqoofxCtaSVGxXr6z1DBIDkpqI5i6JQW'

def get_all_tweets(screen_name, limit=None):
    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_key, access_secret)
    api = tweepy.API(auth)
    	
    alltweets = []
    	
    try:
        new_tweets = api.user_timeline(screen_name = screen_name,count=200)
    except:
        print("Couldn't find tweets for user: {0}".format(screen_name))
        return None
    
    alltweets.extend(new_tweets)
    oldest = alltweets[-1].id - 1
    	
    while len(new_tweets) > 0:
        print "Getting tweets for {0}, {1} tweets dowloaded so far...".format(screen_name, len(alltweets))
        new_tweets = api.user_timeline(screen_name = screen_name,count=200,max_id=oldest)
        alltweets.extend(new_tweets)        
        oldest = alltweets[-1].id - 1
        
        if limit is not None:
            if len(alltweets) > limit:
                alltweets = alltweets[:limit]
                break
            
    print "Finished getting {0} tweets from {1}".format(len(alltweets),screen_name)
     
    return alltweets 
         
def filter_week(alltweets, end_date):
    
    start_date = end_date - datetime.timedelta(days = 7)
    return  [tweet for tweet in alltweets if start_date <= tweet.created_at <= end_date]
         
def create_tweet_sheet(alltweets):

    #outtweets = [[tweet.id_str, tweet.created_at, tweet.text.encode("utf-8")] for tweet in alltweets]
    
    #with open('%s_tweets.csv' % screen_name, 'wb') as f:
    #    writer = csv.writer(f)
    #    writer.writerow(["id","created_at","text"])
    #    writer.writerows(outtweets)
    
    return alltweets

if __name__ == '__main__':
    
    # First lets define a cut off date and user list
    cut_off_date = datetime.datetime(2016,3,2,23,59,59)
    users = ['ConservativesIN','reformineurope','consforbritain','grassroots_out','labour4europe','Scientists4EU','labourleave','StrongerIn','LeaveEUOfficial','vote_leave']
    headings = ['Date','Tweet','Retweets','Favourites','Retweet?','Quoted?','Followers','Following','Retweets-retweeted','favourites-retweeted']
    # Now lets create an excel spreadsheet for the data  
    wb = xl.Workbook()

    # Loop over the users and get their tweets from 7 days before the cutoff
    for user in users:

        # Get all the tweets
        all_tweets = get_all_tweets(user, 500)

        if all_tweets is None:
            continue
        filt_tweets = filter_week(all_tweets, cut_off_date)

        # Create the worksheet and give it a name and add some column heacdings
        ws = wb.create_sheet(0)
        ws.title = user
        
        for row in range(0, len(filt_tweets)+1):

            if row == 0:
                for col, heading in enumerate(headings):
                    wsc = ws.cell(column = col+1, row=row+1)
                    wsc.value = heading
            else:
                text = filt_tweets[row-1].text
                retweet = filt_tweets[row-1].retweeted or text[:2] == u'RT'
                
                wsc = ws.cell(column = 1, row=row+1)
                wsc.value = filt_tweets[row-1].created_at
                
                wsc = ws.cell(column = 2, row=row+1)
                wsc.value = text
                
                wsc = ws.cell(column = 3, row=row+1)
                wsc.value = filt_tweets[row-1].retweet_count
                
                wsc = ws.cell(column = 4, row=row+1)
                wsc.value = filt_tweets[row-1].favorite_count
                
                wsc = ws.cell(column = 5, row=row+1)
                wsc.value = retweet
                
                wsc = ws.cell(column = 6, row=row+1)
                wsc.value = filt_tweets[row-1].is_quote_status
                
                wsc = ws.cell(column = 7, row=row+1)
                wsc.value = filt_tweets[row-1].author.followers_count
                
                wsc = ws.cell(column = 8, row=row+1)
                wsc.value = filt_tweets[row-1].author.friends_count
        
    wb.save("C:\\Users\\kw0020\\TwitterData\\{0}.xlsx".format(cut_off_date.date()))